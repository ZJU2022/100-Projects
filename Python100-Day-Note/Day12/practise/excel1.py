#创建excel文件
#https://openpyxl-chinese-docs.readthedocs.io/zh_CN/latest/tutorial.html 文档

#无须在文件系统中创建文件即可开始使用openpyxl。只要导入 Workbook 类就可以开始工作了:
from openpyxl import Workbook,worksheet
wb = Workbook()
#一个工作表至少有一个工作簿. 你可以通过 Workbook.active 来获取这个属性:
ws = wb.active
#你可以使用 Workbook.create_sheet 方法来创建新的工作簿:
ws1 = Workbook.create_sheet('MySheet')  #insert at the end
ws2 = Workbook.create_sheet("Mysheet1",0) #insert at first position
ws3 = wb.create_sheet("Mysheet2", -1) # insert at the penultimate position

#工作薄在创建时会自动生成一个名字，以(Sheet, Sheet1, Sheet2, …)来进行命名。你也可以通过 Worksheet.title 属性来修改命名:
ws.title('linfeiwu_Sheet')
#默认情况下，包含该标题的选项卡的背景颜色为白色。你也可以使用 RRGGBB 颜色来改变 Worksheet.sheet_properties.tabColor 属性:
ws.sheet_properties.tabColor = "1072BA"
#给工作表命名后，就可以将其作为工作簿点键
ws3 = wb("linfeiwu_Sheet")
#你可以使用 Workbook.sheetname 属性查看工作簿中所有工作表的名称:
print(wb.sheetnames)
#可以遍历工作表
for sheet in wb:
    print(sheet.title)

'''
新建工作表
访问单元格/访问大量单元格/数据存储（保存至文件/保存成流/从文件加载）
写入工作表
读取工作表
合并/拆分单元格
使用数字格式/公式
插入图像，隐藏
。。。。。。。
'''

